##!/usr/bin/python


# https://noisefloor-net.blogspot.com/2015/07/python-tabellen-in-reportlab.html

from reportlab.platypus import Paragraph, PageBreak, Table, Image,TableStyle, SimpleDocTemplate, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.graphics import renderPDF
from reportlab.lib import colors
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.barcharts import HorizontalBarChart
from reportlab.graphics.shapes import Drawing, _DrawingEditorMixin, String
from reportlab.lib.colors import blue
from reportlab.graphics.charts.textlabels import Label
from reportlab.graphics.charts.legends import Legend
import matplotlib.pyplot as plt
from reportlab.lib.units import cm,inch
import copy
from collections import Counter
import numpy as np
import zipfile
from openpyxl import load_workbook #used for EXCEL Sheet
from openpyxl import Workbook #used for EXCEL Sheet
import re
import requests
import pandas as pd
import os
#from bs4 import BeautifulSoup #   http://www.crummy.com/software/BeautifulSoup/#Download
#from openpyxl.cell import get_column_letter
#https://www.zewaren.net/reportlab.html
#https://developers.google.com/kml/documentation/kmzarchives wichtige infos zu KMZ Dateien
#    conda activate spyder-env

hauptdatei="Teufelsturmauswertung/Hauptversion_Elbi.xlsx"
## Dateien laden
wb = load_workbook(filename = hauptdatei)
#wb = load_workbook(filename = 'write2cell Kopie.xlsx')
#wb = load_workbook(filename =hauptdatei)

ws = wb.active
wb.title = "Entfernungen"
GEBIET_AUSWAHL="Rathen"  # fertig 11.6.23
#GEBIET_AUSWAHL = 'Schrammsteine'
#GEBIET_AUSWAHL = 'Schmilka'
#GEBIET_AUSWAHL = 'Affensteine' #fertig
#GEBIET_AUSWAHL = 'Alle'
#GEBIET_AUSWAHL = 'Brandgebiet' gibts nicht
#GEBIET_AUSWAHL = 'Zschand' Fertig 11.6.23

#GEBIET_AUSWAHL = 'Bielatal' #fertig
#GEBIET_AUSWAHL = wb['Test']
GEBIET_AUSWAHL = 'Gebiet der Steine' ##fertig 11.6.23
sheet_ranges = wb[GEBIET_AUSWAHL] # fertig 11.6.23
Ausgabe_Name=GEBIET_AUSWAHL+".kml"

#sheet_ranges = wb['Elbi']
def extrahiere_grad(text):
    if isinstance(text,int):
        return text
    else:   

        text=text.replace("RP ","")
        text=text.replace("! ","")
        text=text.replace("!","")
        text=text.replace("** ","")
        text=text.replace("* ","")
        text=text.replace("(","")
        text=text.replace(")","")
        tabelle=text.split(" ")
        return tabelle[0]

def Grade_zu_Histo_Array(Histo_Dict,schwierigkeitsgrade):
    ausgabe=np.zeros(len(schwierigkeitsgrade))
    for Grad in Histo_Dict:
        Anzahl=Histo_Dict[Grad]
        Grad=Grad.replace("RP ","")
        Grad=Grad.replace("! ","")
        Grad=Grad.replace("!","")
        Grad=Grad.replace("* ","")
        Grad=Grad.replace("** ","")
        Grad=Grad.split(" ")[0]
        
        i=0
        while i<len(schwierigkeitsgrade):
            if schwierigkeitsgrade[i]==Grad:
                ausgabe[i]=Anzahl
            i=i+1
        
        print(Grad,Anzahl)
    print(ausgabe)
    return ausgabe

def wert_zu_zahl(text):
    if isinstance(text,str):
   
        buchstabe=0
        
        if re.findall("a",text):
            buchstabe=-0.3
        if re.findall("b",text):
            buchstabe=0
        if re.findall("c",text):
            buchstabe=0.3
     
        
        if re.findall("VIII",text):
            return 8+buchstabe    
        
        if re.findall("VII",text):
            return 7+buchstabe      
        
        if re.findall("VI",text):
            return 6+buchstabe      
        
        if re.findall("IV",text):
            return 4+buchstabe     
       
        if re.findall("V",text):
            return 5+buchstabe      
                
        if re.findall("XI",text):
            return 11+buchstabe    
            
        if re.findall("IX",text):
            return 9+buchstabe
        
        if re.findall("X",text):
            return 10+buchstabe
        
        if re.findall("III",text):
            return 3+buchstabe  
        
        if re.findall("II",text):
            return 2+buchstabe      
    
        if re.findall("I",text):
            return 1+buchstabe  
        
    else:
        return text

def umlaute_tauschen(string):
    string=string.replace("ä","ae")
    string=string.replace("ö","oe")
    string=string.replace("ü","ue")
    string=string.replace("ß","ss")
    return string
                   

def write_logs(file,line):
        #write out the event, open and close the file each time for proper tailing
        #Time= datetime.datetime.fromtimestamp(int(Zeitpunkt)).strftime('%Y-%m-%d %H:%M:%S')
        output_file = open(file, 'a')
        for column in line:
            output_file.write(str(column)+";")
        output_file.write(str("\n"))
        output_file.close()



# KML Text anpassen mit PDF und Icons
def KML_TEXT_SCHREIBEN(name,laenge,breite,hoehe,farbe):
    pfad     =umlaute_tauschen(name)+".pdf"
    pfad_png =umlaute_tauschen(name)+".png"
    pfad_html=umlaute_tauschen(name)+".html"
    print("Pfad",pfad,name)
    string="<Placemark><name>"+name+"</name>\n"+"<Style><IconStyle><Icon><href>files/"+farbe+".png</href></Icon></IconStyle></Style>\n"+"<ExtendedData>\n"+"<lc:attachment>files/"+pfad_html+"</lc:attachment>\n"+"<lc:attachment>files/"+pfad_png+"</lc:attachment>\n"+"<lc:attachment>files/"+pfad+"</lc:attachment>\n"+"</ExtendedData>\n"+">	<Point>\n"+"<coordinates>"+str(breite)+","+str(laenge)+","+str(hoehe)+"</coordinates>\n"+"</Point>" +"</Placemark>"    
    return string
    

# Aus den Pfeilen Werte generien
def umbennung(arr):
    eintrag=[]
    for wert in arr:
        if "arrow-upright3.gif"==wert:
            eintrag.append(1)
        if "arrow-upright2.gif"==wert:
            eintrag.append(2)
        if "arrow-upright.gif" ==wert:
            eintrag.append(3)
        if "arrow-right.gif"   ==wert:
            eintrag.append(4)
        if "arrow-downright.gif"==wert:
            eintrag.append(5)
        if "arrow-downright2.gif"==wert:
            eintrag.append(6)
        if "arrow-downright3.gif"==wert:
            eintrag.append(7)
        #else:
        #    eintrag.append(0)

    return eintrag


#Extrahiere die Wegnummern der einzelnen Wege
    #https://www.teufelsturm.de/wege/bewertungen/anzeige.php?wegnr=9847
def extrahiere_Wegnummer(Seiten_Text):
    Weg_Nummer_Links=re.findall("wege/bewertungen/anzeige.php\?wegnr=\d+",Seiten_Text)  
    #print(Weg_Nummer_Links)
    Weg_Nummer=[]
    for Wert in Weg_Nummer_Links:
        Weg_Nummer.append(int(re.findall("\d+",Wert)[0]))   
    return Weg_Nummer



# Informationen zu den Wegen
def WEGE_AUSLESEN(Gipfelnummer,Gipfel,Gebiet,file_tabelle):
    hauptseite='https://www.teufelsturm.de/wege/suche.php?gipfelnr='+str(Gipfelnummer)
    print("HAUPTSEITE",hauptseite)
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

    r = requests.get(hauptseite,headers=headers)
    df = pd.read_html(r.text) # this parses all the tables in webpages to a list
    Pfeil_Bewertung=re.findall("arrow-\S+.gif",r.text)
    Pfeil_Bewertung=umbennung(Pfeil_Bewertung) # Aus Pfeilicons Wertungen machen
    
    #Wege auslesen
    Wege_Nummer=extrahiere_Wegnummer(r.text)
    Wege_Nummer=Wege_Nummer[::2]
    print("Wege_Nummer",Wege_Nummer)
    
    Ausgabe_Liste=[["Nummer",'Routenname', 'Bewertung', 'Schwierigkeit']] #Header
    Wege_Name=[]
    print ("Anzahl_Routen",len(df[4])-1, "Anzahl Einträge Tabelle",len(df[4]))
    print ("Anzahl Wege Nummer",len(Wege_Nummer))
    print ("Anzahl Pfeil Bewertung",len(Pfeil_Bewertung))
    
    if len(Pfeil_Bewertung)!=len(Wege_Nummer):
        Pfeil_Bewertung.append(0)    
        

    for i in range(1,len(df[4])): # jede Zeile durchgehen
        print(i,"Weg:",df[4][2][i],df[4][0][i],df[4][2][i] , Pfeil_Bewertung[i-1], df[4][4][i],"Nummer:",Wege_Nummer[i-1] )
        #print(i,"Mittlere Zahl:   ",df[4][i][0])
        Wege_Name.append(df[4][2][i])
        Ausgabe_Liste.append([df[4][0][i],df[4][2][i] , Pfeil_Bewertung[i-1], df[4][4][i]  ])
        text=extrahiere_grad(df[4][4][i])      
        zahl=wert_zu_zahl(text)
        write_logs(file_tabelle,[Gebiet,Gipfel,df[4][0][i],df[4][2][i] , Pfeil_Bewertung[i-1], df[4][4][i], zahl  ])
    return Ausgabe_Liste,Wege_Nummer,Wege_Name


# Kommentare der einzelnen Wege
def WEGE_KOMMENTARE(Weg_Nummer):
    hauptseite="https://www.teufelsturm.de/wege/bewertungen/anzeige.php?wegnr="+str(Weg_Nummer)
    #response = urllib3.urlopen(hauptseite)
   # print("Seite für Routenauswertung:",hauptseite)   
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
     
    r = requests.get(hauptseite,headers=headers)
    df_list = pd.read_html(r.text) # this parses all the tables in webpages to a list
    df = df_list[-1] #3
    Ausgabe_Liste=[["Nutzer","Kommentar","Bewertung"]]
    for i in range(1,len(df[0])): # jede Zeile durchgehen
        """
        print(i,"Bewertung:   ",df[2][i])
        print(i,"Kommentar:   ",df[1][i])
        print(i,"Nutzer:      ",df[0][i])
        print("----------")
        """
        Ausgabe_Liste.append([df[0][i],df[1][i],df[2][i]])

    return Ausgabe_Liste



def Anzahl_Benotungen(schwierigkeiten,benotungen,bewertungs_skala):
    # erstellt aus dem Schwierigkeiten und Benotungen eine Liste mit der Anzahl
    skala_zahlen=[i for i in range(len(bewertungs_skala))]
    noten_skala =[1,2,3,4,5,6,7]
    schwierigkeiten=[extrahiere_grad(i) for i in schwierigkeiten] 
    print("schwierigkeiten Extrahiert",schwierigkeiten)
    schwierigkeiten_zahlen=[]
    for i in range(len(schwierigkeiten)):
        try:
            schwierigkeiten_zahlen.append(bewertungs_skala.index(schwierigkeiten[i]))
        except:
            schwierigkeiten_zahlen.append(6)
    matrix_anzahl=np.zeros((len(bewertungs_skala),len(noten_skala)))
    matrix_noten=copy.deepcopy(matrix_anzahl)
    matrix_grad_zwi=[[bewertungs_skala[i]]*7 for i in range(len(bewertungs_skala))]
    matrix_grad=[]
    for sub in matrix_grad_zwi:
        for i in sub:
            matrix_grad.append(i)       
    # Matrix mit Anzahl befüllen        
    i=0
    while i< len(schwierigkeiten):   
        #print(i,schwierigkeiten_zahlen[i],benotungen[i])
        matrix_anzahl[schwierigkeiten_zahlen[i],benotungen[i]-1]=matrix_anzahl[schwierigkeiten_zahlen[i],benotungen[i]-1]+1
        i=i+1
    
    for skala_i in range(len(bewertungs_skala)):
        i=0
        for noten_i in range(len(noten_skala)):  
            #print("Noten i",noten_i,"Skala i",skala_i)
            matrix_noten[skala_i][noten_i]=noten_skala[noten_i]
            i=i+1
               
    return matrix_anzahl.flatten(),matrix_noten.flatten(),matrix_grad




def PLOT_Schwierigkeit_vs_Benotung(schwierigkeiten_org,benotung,ordner,gipfel,skala):
        
    anzahl,benotung,schwierigkeiten=Anzahl_Benotungen(schwierigkeiten_org,benotung,skala)
    
    #print("Anzahl",anzahl)
    #print("Benotung",benotung)
    #print("Schwierigkeiten",schwierigkeiten)
    
    anzahl_innen =np.array(anzahl)*200
    anzahl_aussen=anzahl_innen*0.9
    
    fig, ax3 = plt.subplots(figsize=(8,6))
    ax3.set_title(gipfel[6:])

    ax3.tick_params(axis='x', which='major', labelsize=10)
    ax3.tick_params(axis='y', which='major', labelsize=12)
    ax3.grid(zorder=0)
    ax3.set_yticks([1,2,3,4,5,6,7])
    ax3.set_ylim((0.1,7.9))
    ax3.set_yticklabels(["+++","++","+","o","-","--","---"])

    ax3.set_xlim((-1,len(skala)))
      
    ax3.scatter(schwierigkeiten[:], benotung[:], s=anzahl_innen,  alpha=1,zorder=3,color="k")
    ax3.scatter(schwierigkeiten[:], benotung[:], s=anzahl_aussen, alpha=1,zorder=3,color="dodgerblue")
    #ax3.text(0,1,"1",horizontalalignment='center',verticalalignment='center',zorder=4)
    
    for i in range(len(schwierigkeiten)):
        if anzahl[i]>0:
            ax3.text(schwierigkeiten[i],benotung[i],int(anzahl[i]),horizontalalignment='center',verticalalignment='center',zorder=4)   
    
    ax3.set_xlabel("Schwierigkeit")
    ax3.set_ylabel("Bewertung")
    plt.savefig(ordner+gipfel+".png",dpi=150)
    return anzahl,benotung,schwierigkeiten



def erstelle_KMZ(Gipfel_Array,Ausgabe_Name):
     output=Ausgabe_Name[:-3]+"_Zusammen.kmz"    
     output="Zusammen.kmz"     
 
     zf = zipfile.ZipFile(output,'a')      
     for gipfel in Gipfel_Array:
         try:
             image="files/"+gipfel+".pdf"
             image_umlaute="files/"+umlaute_tauschen(gipfel)+".pdf"
             zf.write(image,arcname=image_umlaute) ##Relative Path to the Image 
             image="files/"+gipfel+".png"
             image_umlaute="files/"+umlaute_tauschen(gipfel)+".png"
             zf.write(image,arcname=image_umlaute) ##Relative Path to the Image             
             image="maps/"+gipfel+".html"
             image_umlaute="files/"+umlaute_tauschen(gipfel)+".html"
             try:
                 zf.write(image,arcname=image_umlaute) ##Relative Path to the Image   
             except:
                 print("Konnte Karte nicht finden")
             print(image_umlaute)
         except:
             print("PDF nicht gefunden",gipfel)
     #zf.write("icon_default.png",arcname="files/icon_default.png") ##Relative Path to the Image 
     zf.write("Icons/green.png", arcname="files/green.png") ##Relative Path to the Image 
     zf.write("Icons/orange.png",arcname="files/orange.png") 
     zf.write("Icons/red.png",   arcname="files/red.png") 
     zf.write("Icons/blue.png",  arcname="files/blue.png") 
     zf.write(Ausgabe_Name,arcname=Ausgabe_Name) ##Add revised doc.kml file 
     
     zf.close() 
    


def entferne_spruenge(schwierigkeiten,benotungen):
    out_schwie,out_benot=[],[]
    for i in range(len(schwierigkeiten)):   
        if schwierigkeiten[i].find("1")>=0 or schwierigkeiten[i].find("2")>=0 or schwierigkeiten[i].find("3")>=0 or schwierigkeiten[i].find("4")>=0  or schwierigkeiten[i].find("5")>=0:
            print("Sprung gefunden und entfernt",schwierigkeiten[i])     
        elif not schwierigkeiten[i].isdigit():
            out_schwie.append(schwierigkeiten[i])
            out_benot.append(benotungen[i])        
        else:
            print("Sprung gefunden und entfernt")
    return out_schwie,out_benot
    

def bewertungs_kategorie(anzahl,benotung,schwierigkeiten):
    schwierigkeiten_als_zahl=[wert_zu_zahl(i) for i in schwierigkeiten]
    counter=0
   #print("schwierigkeiten_als_zahl[i],  benotung[i], anzahl[i]")
    max_schwierigkeit=6
    max_benotung=4

    for i in range(len(schwierigkeiten)):
        if (schwierigkeiten_als_zahl[i]<=max_schwierigkeit) and benotung[i]<max_benotung and anzahl[i]>0:
            counter=counter+anzahl[i]*(max_benotung-benotung[i])
            #print(schwierigkeiten_als_zahl[i],  benotung[i], anzahl[i]) 
    if counter<3:
        wert= "red"   
    elif counter<6:
        wert= "orange"    
    elif counter<10:
        wert= "green"   
    elif counter>=10:
        wert= "blue"
    print("Iconfarbe: ",wert, " mit Counter: ",counter)  
    return wert


 
file=open("ROH.KML")
file_ausgabe=open("AUSGABE.KML","w+")
file_tabelle="Tabelle_Wege.csv"
tabelle=open(file_tabelle,"w+")
tabelle.write("Gebiet;Gipfel;Nummer;Wegname;Bewertung;Schwierigkeit;Schwierigkeit als Zahl;\n")
tabelle.close()


Zeilen=file.readlines()
file.close()

#Kopiere den ROH Text in AUSGABE
for Zeile in Zeilen:
    file_ausgabe.write(Zeile)

#os.system('rm "/Users/nicefuchs/Desktop/Klettern/Kletter Programme/GPS mit PDF Projekt/files/Sheriff.pdf" -f')
#os.system('rm "/Users/nicefuchs/Desktop/Klettern/Kletter Programme/GPS mit PDF Projekt/files/Herkulesstein.pdf" -f')

PDF_Ordner=os.listdir("files")
PDF_Liste_Str="; ".join(PDF_Ordner)
#print("Vorhandene PDFs",PDF_Liste_Str)


schwierigkeits_skala=["I","II","III","IV","V","VI","VIIa","VIIb","VIIc","VIIIa","VIIIb","VIIIc","IXa","IXb","IXc","Xa","Xb","Xc"]




i=2 #i=2 Standard
Gipfel_Array=[]
i_max=42000 #hohe zahl falls alle, kleine zahl bei Tests
while sheet_ranges['A'+str(i)].value!=None and i<i_max:
    print("\n\n\n--------------------------------------------------")
    print("---------------------NEUER GIPFEL-----------------------------")
    print("--------------------------------------------------")


    # Gipfel Koordinaten einlesen aus EXCEL
    GIPFEL_NUMMER=str((sheet_ranges['H'+str(i)].value)) # die von der URL ist wichtig
    GIPFEL=(sheet_ranges['B'+str(i)].value) # B
    gebiet=(sheet_ranges['D'+str(i)].value) # D
    laenge=str(sheet_ranges['F'+str(i)].value) # D
    breite=str(sheet_ranges['G'+str(i)].value) # E 
    
    print ("i",i,"Laenge:",laenge,"Breite:",breite,"Gipfel_Name",GIPFEL,"Gipfelnummer",GIPFEL_NUMMER)


    if laenge!="0":   
        try:
            distanz_1=int(sheet_ranges['I'+str(i)].value) # I
            dauer_1  =int(sheet_ranges['J'+str(i)].value) # J 
            parkplatz_1=str(sheet_ranges['I1'].value) # J 
            info_1=[parkplatz_1,distanz_1,dauer_1,"Blau"]
            
            distanz_2=int(sheet_ranges['K'+str(i)].value) # I
            dauer_2  =int(sheet_ranges['L'+str(i)].value) # J 
            parkplatz_2=str(sheet_ranges['K1'].value) # J 
            info_2=[parkplatz_2,distanz_2,dauer_2,"Rot"] 
        
            distanz_3=int(sheet_ranges['M'+str(i)].value) # I
            dauer_3  =int(sheet_ranges['N'+str(i)].value) # J 
            parkplatz_3=str(sheet_ranges['M1'].value) # J 
            info_3=[parkplatz_3,distanz_3,dauer_3,"Orange"]    
        except:
            distanz_1,distanz_2,distanz_3="?","?","?"
            dauer_1,dauer_2,dauer_3="?","?","?"
            parkplatz_1,parkplatz_2,parkplatz_3="?","?","?"
            info_1=[parkplatz_1,distanz_1,dauer_1,"Blau"]
            info_2=[parkplatz_2,distanz_2,dauer_2,"Rot"] 
            info_3=[parkplatz_3,distanz_3,dauer_3,"Orange"]  

    else:
        distanz_1,distanz_2,distanz_3="?","?","?"
        dauer_1,dauer_2,dauer_3="?","?","?"
        parkplatz_1,parkplatz_2,parkplatz_3="?","?","?"
        info_1=[parkplatz_1,distanz_1,dauer_1,"Blau"]
        info_2=[parkplatz_2,distanz_2,dauer_2,"Rot"] 
        info_3=[parkplatz_3,distanz_3,dauer_3,"Orange"]    


    print("1: distanz",distanz_1,"Dauer",dauer_1)
    print("2: distanz",distanz_2,"Dauer",dauer_2)
    print("3: distanz",distanz_3,"Dauer",dauer_3)

    
    Gipfel_Array.append(GIPFEL)
    
    i=i+1


    # Gibt Infos zu Wegen wieder
    #try:
    #if os.path.isfile("files/"+GIPFEL+".pdf")==False:
    try:   
        Weg_Array,Weg_Nummer,Weg_Namen = WEGE_AUSLESEN(GIPFEL_NUMMER,GIPFEL,gebiet,file_tabelle)
    except:
        Weg_Array,Weg_Nummer,Weg_Namen = ["-"],[-1],["Nichts gefunden"]

    print("Weg_Nummern",Weg_Nummer)
    print("Weg_Namen",Weg_Namen)
    
            
    # Aus Noten wieder +++--- machen
    Weg_Array_Pluszeichen=copy.deepcopy(Weg_Array)
    Plus_Minus=["+++","++","+","o","-","--","---"]        
    for p in range(1,len(Weg_Array_Pluszeichen)):
        Note_zwi=int(Weg_Array_Pluszeichen[p][2])-1
        Weg_Array_Pluszeichen[p][2]=Plus_Minus[Note_zwi]
                    
  
    # Daten für Auswertung
    Noten=    [i[2] for i in Weg_Array[1:]]            
    Bewertung=[i[3] for i in Weg_Array[1:]]
    
    Bewertung,Noten=entferne_spruenge(Bewertung,Noten)
    print("Bewertung:",Bewertung)
    print("Noten",Noten)
    
    #Erstelle den ScatterPlot
    anzahl,benotung,schwierigkeiten=PLOT_Schwierigkeit_vs_Benotung(Bewertung,Noten,"Scatterplot/",GIPFEL,schwierigkeits_skala)
    
    #lege Farbe fest für Icon
    farbe=bewertungs_kategorie(anzahl,benotung,schwierigkeiten)  
  
    Gipfel_Dot_PDF=str(GIPFEL)+'.pdf'
    #if len(re.findall(Gipfel_Dot_PDF,PDF_Liste_Str))==0: #falls PDF existiert
    if os.path.isfile("files/"+GIPFEL+".pdf")==False:
        print("ERZEUGE NEUES PDF----------------------",GIPFEL)

    
        # PDF mit Infos schreiben
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate("PDF_Ausgabe/"+Gipfel_Dot_PDF, pagesize=A4,rightMargin=10,leftMargin=10, topMargin=30,bottomMargin=18)
        story = []
        Stil_Paragraph=styles['Normal']
        Stil_Paragraph.fontSize=15
        story.append(Paragraph('Routen von Gipfel '+str(GIPFEL)+ " im Gebiet "+str(gebiet), styles['Normal']))           
        t = Table(Weg_Array_Pluszeichen)
        t.hAlign = 'LEFT'
        t.spaceBefore =  10
        t.spaceAfter = 10
        t.setStyle(TableStyle(
            [('BOX', (0,0), (-1,-1), 0.5, colors.black),
             ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black)]))           
        story.append(t)
        #story.append("Die Infos stammen von www.teufelsturm.de")

        # Anreise Infos
        story.append(Paragraph('Distanz von Haltestelle/Parkplatz zu Gipfel', styles['Normal']))
        t = Table([ ["Ort","Distanz(m)","Zeit(min)","Farbe Karte"],info_1,info_2,info_3])
        t.hAlign = 'LEFT'
        t.spaceBefore =  10
        t.spaceAfter = 10
        t.setStyle(TableStyle(
            [('BOX', (0,0), (-1,-1), 0.5, colors.black),
             ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black)]))           
        story.append(t)

    
        
        story.append(PageBreak())
        story.append(Image("Scatterplot/"+GIPFEL+".png"    ,width=2.2*8*cm, height=2.2*6*cm))
        if os.path.exists("maps/"+GIPFEL+"_map.png"):          
            story.append(Image("maps/"+GIPFEL+"_map.png",width=2.2*8*cm, height=2.2*6*cm))
        else:
            print("Keine Karte vorhanden")
        story.append(PageBreak())
        
        for Weg_Zahler in range(len(Weg_Nummer)):
        #for Weg_Zahler in range(2):
            #if True:
            try:
                #print("-------------------------------------")
                #Gibt die Kommentare der Wege wieder
                Kommentare_Wege=WEGE_KOMMENTARE(Weg_Nummer[Weg_Zahler])
                print(Weg_Zahler,"Weg_Zahler",Weg_Namen[Weg_Zahler],"Weg:",Weg_Nummer[Weg_Zahler])
                story.append(Paragraph("Wegname:"+Weg_Namen[Weg_Zahler]+" ("+Bewertung[Weg_Zahler]+")", Stil_Paragraph))
    
                #print(Kommentare_Wege)
                
                s = getSampleStyleSheet()
                s = s["BodyText"]
                s.wordWrap = 'CJK' #für Zeilenbruch
                data2 = [[Paragraph(cell, s) for cell in row] for row in Kommentare_Wege]
                t=Table(data2)
                t = Table(data2,colWidths=(50,420,70))
                t.hAlign = 'LEFT'
                t.spaceBefore =  10
                t.spaceAfter = 10                
                t.wordWrap = 'CJK'
                t.setStyle(TableStyle(
                    [('BOX', (0,0), (-1,-1), 0.5, colors.black),
                     ('VALIGN',(0,0),(-1,-1),'TOP'),
                     ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black)]))
                story.append(t)
                
    
                #print (WEGE_KOMMENTARE(Wege))
            except:
                print ("fehler bei Kommentaren")   
    
        doc.build(story)
    else:
        print("PDF Existiert Bereits:",GIPFEL)  


    hoehe =100.0
    
    KML_Text=KML_TEXT_SCHREIBEN(GIPFEL,laenge,breite,hoehe,farbe)
    file_ausgabe.write(KML_Text+"\n\n\n")

    #except:
        #print("Keine Wege gefunden")

file_ausgabe.write("</Document></kml>")
file_ausgabe.close()

erstelle_KMZ(Gipfel_Array,Ausgabe_Name) 
