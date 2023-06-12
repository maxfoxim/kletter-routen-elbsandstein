import time
from selenium import webdriver
import openrouteservice
import re
from openpyxl import load_workbook #used for EXCEL Sheet
from openpyxl import Workbook #used for EXCEL Sheet
import folium
from shapely.geometry import LineString, Polygon, mapping
from shapely.ops import cascaded_union
from PIL import Image
import json
import os.path
#wb = load_workbook(filename = 'MinimalElbi.xlsx')
#wb = load_workbook(filename = 'Test_Excel.xlsx')

#hauptdatei="MinimalElbi.xlsx"
hauptdatei="Teufelsturmauswertung/Hauptversion_Elbi.xlsx"

wb = load_workbook(filename = hauptdatei)

ws = wb.active
wb.title = "Entfernungen"
#sheets=["Gebiet der Steine"]
#GEBIET_AUSWAHL = "Rathen"  # fertig 11.6.23
#GEBIET_AUSWAHL = 'Schrammsteine' fehlen Parkplätze
#GEBIET_AUSWAHL = 'Schmilka' fehlen Parkplätze
#GEBIET_AUSWAHL = 'Affensteine' #fertig 7.6.21
#GEBIET_AUSWAHL = 'Brandgebiet' gibts nicht als eigenen Reiter, fehlen Parkplätze
#GEBIET_AUSWAHL = 'Zschand' Fertig 11.6.23
#GEBIET_AUSWAHL = 'Bielatal' #fertig 16.5.21
GEBIET_AUSWAHL = 'Gebiet der Steine' ##fertig 11.6.23
sheets = [GEBIET_AUSWAHL] # fertig 11.6.23


"""
Weitere Ideen:
    -Map mit Gipfelqualtität aller Gipfel
"""

def style_function(color): # To style data
    return lambda feature: dict(color=color,
                                opacity=0.9,
                                weight=4,)


def suche_koordinaten(text):
    p=re.compile("\d+\.\d+") #REGEX
    koor=p.findall(text)
    print(koor)
    return (float(koor[1]),float(koor[0]))
    


def distanzen_berechnen(Start,Ende,datei):
    coords=(Start,Ende)
    client = openrouteservice.Client(key='HIER Schlüssel eintragen von OPEN ROUTE Service') # Specify your personal API key
    ausgabe = client.directions(coords,profile='foot-walking',geometry= 'true',format_out="geojson")
    #print(ausgabe)
    distanz=ausgabe["features"][0]["properties"]["summary"]["distance"]
    dauer=ausgabe["features"][0]["properties"]["summary"]["duration"]/60.
    bbox=ausgabe["features"][0]["bbox"]
    print("Distanz",distanz,"Dauer",dauer)
    with open(datei, 'w') as outfile:
        json.dump(ausgabe, outfile)
    return int(distanz), int(dauer),ausgabe 

def distanzen_aus_json(ausgabe):
    distanz=ausgabe["features"][0]["properties"]["summary"]["distance"]
    dauer=ausgabe["features"][0]["properties"]["summary"]["duration"]/60.
    return distanz,dauer


def berechne_karten_ausschnitt(lon_arr,lat_arr):
    differenz=0.0025 # vergrößerung damit alles rein passt
    lon_min=min(lon_arr)-differenz # links
    lon_max=max(lon_arr)+differenz # rechts
    lat_min=min(lat_arr)-differenz # unten
    lat_max=max(lat_arr)+differenz # oben
    return [(lat_min,lon_min),(lat_max,lon_max)]


def route_darstellen_als_html(Gipfel,routen_GPS_koor_1,routen_GPS_koor_2,routen_GPS_koor_3,Start_1,Start_2,Start_3,ziel):    
    bbox_ausschnitt=berechne_karten_ausschnitt([Start_1[0],Start_2[0],Start_3[0],ziel[0]],[Start_1[1],Start_2[1],Start_3[1],ziel[1]])
    print(bbox_ausschnitt)
    kartengrundlage = folium.Map(location=(ziel), zoom_start=16) # Create map
    kartengrundlage.fit_bounds(bbox_ausschnitt)
    
    #Zeichne Routen
    folium.GeoJson(routen_GPS_koor_1,name='1 Route',style_function=style_function('blue'))  .add_to(kartengrundlage)
    folium.GeoJson(routen_GPS_koor_2,name='2 Route',style_function=style_function('red'))   .add_to(kartengrundlage)
    folium.GeoJson(routen_GPS_koor_3,name='3 Route',style_function=style_function('orange')).add_to(kartengrundlage)

    #Setze Markierungen für Startpunkte
    folium.Marker(list(reversed(Start_1)), popup='Start_1',icon=folium.Icon(color='white',icon_color='blue',icon='taxi',prefix='fa')).add_to(kartengrundlage)
    folium.Marker(list(reversed(Start_2)), popup='Start_2',icon=folium.Icon(color='white',icon_color='red',icon='taxi',prefix='fa')).add_to(kartengrundlage)
    folium.Marker(list(reversed(Start_3)), popup='Start_3',icon=folium.Icon(color='white',icon_color='orange',icon='taxi',prefix='fa')).add_to(kartengrundlage)
    
    # Zeichne Gipfel
    folium.Marker(list(reversed(ziel )),   popup='Ziel') .add_to(kartengrundlage)
    kartengrundlage.save("maps/"+Gipfel+".html")


def HTML_TO_PNG(Gipfel):
    driver = webdriver.Chrome(executable_path="/Users/stephan/Library/CloudStorage/OneDrive-Persönlich/Klettern/chromedriver")
    driver.set_window_size(1000, 800)  # choose a resolution
    driver.get("file:////Users/stephan/Library/CloudStorage/OneDrive-Persönlich/Klettern/maps/"+Gipfel+".html")
    time.sleep(2)
    driver.save_screenshot("maps/"+Gipfel+"_largemap.png")        
    driver.quit()
    #Bild verkleinern
    image = Image.open("maps/"+Gipfel+"_largemap.png")
    bild_groesse=1100
    new_image = image.resize((bild_groesse, int(bild_groesse*1354./2000.)))
    new_image.save("maps/"+Gipfel+"_map.png")


col=9
row=2
ROW_MAX=5000 # hohe zahl falls alle, kleine zahl bei Tests
Wartezeit=0.01 # Wartezeit da ein Limit bei der API besteht
routen_nicht_neu_berechnen=False

for sheet in sheets:  
    print("##############################")
    print("SHEET:",sheet)
    ws = wb[sheet]
    
    while ws['A'+str(row)].value!=None:
        
        if row<ROW_MAX:

            GIPFEL=(ws['B'+str(row)].value) 
            print("----------------------")
            print("Reihe",row,"Gipfel",GIPFEL)
            
            Start_1=suche_koordinaten(ws['I1'].value)
            Start_2=suche_koordinaten(ws['K1'].value)
            Start_3=suche_koordinaten(ws['M1'].value)
    
            #Gipfel Koordinaten
            laenge=str(ws['F'+str(row)].value) # D
            breite=str(ws['G'+str(row)].value)
            if breite!="0":
                
                Ende=(float(breite),float(laenge))
                #time.sleep(1)
                
                if os.path.isfile("routen_json/"+GIPFEL+"_1.json") and routen_nicht_neu_berechnen:
                    print("Daten werden aus JSON gelesen")
                    with open("routen_json/"+GIPFEL+"_1.json") as f:
                        routen_GPS_koor_1 = json.load(f)
                    with open("routen_json/"+GIPFEL+"_2.json") as f:
                        routen_GPS_koor_2 = json.load(f)     
                    with open("routen_json/"+GIPFEL+"_3.json") as f:
                        routen_GPS_koor_3 = json.load(f)  
                        
                    distanz_1, dauer_1=distanzen_aus_json(routen_GPS_koor_1)
                    distanz_2, dauer_2=distanzen_aus_json(routen_GPS_koor_2)
                    distanz_3, dauer_3=distanzen_aus_json(routen_GPS_koor_3)
                    
                else:
                    print("Route neu berechnen")
                    #Berechne Route
                    try:
                        distanz_1, dauer_1,routen_GPS_koor_1=distanzen_berechnen(Start_1,Ende,"routen_json/"+GIPFEL+"_1.json")
                        distanz_2, dauer_2,routen_GPS_koor_2=distanzen_berechnen(Start_2,Ende,"routen_json/"+GIPFEL+"_2.json")
                        distanz_3, dauer_3,routen_GPS_koor_3=distanzen_berechnen(Start_3,Ende,"routen_json/"+GIPFEL+"_3.json")
                        time.sleep(Wartezeit)
                                                                    
                        route_darstellen_als_html(GIPFEL,routen_GPS_koor_1,routen_GPS_koor_2,routen_GPS_koor_3,Start_1,Start_2,Start_3,Ende)

                    except:
                        distanz_1, dauer_1,routen_GPS_koor_1="?","?","?"
                        distanz_2, dauer_2,routen_GPS_koor_2="?","?","?"
                        distanz_3, dauer_3,routen_GPS_koor_3="?","?","?"



                #Trage Werte in Excel ein
                ws.cell(column=col,   row=row,   value=distanz_1) #lat
                ws.cell(column=col+1, row=row,   value=dauer_1) #lat
                ws.cell(column=col+2, row=row,   value=distanz_2) #lat
                ws.cell(column=col+3, row=row,   value=dauer_2) #lat
                ws.cell(column=col+4, row=row,   value=distanz_3) #lat
                ws.cell(column=col+5, row=row,   value=dauer_3) #lat                    
                    
                            
                
                HTML_TO_PNG(GIPFEL)                          

                        

              
                    
            else:    
                distanz_1,distanz_2,distanz_3="?","?","?"
                dauer_1,dauer_2,dauer_3="?","?","?"
                               
                ws.cell(column=col,   row=row,   value=distanz_1) #lat
                ws.cell(column=col+1, row=row,   value=dauer_1) #lat
                ws.cell(column=col+2, row=row,   value=distanz_2) #lat
                ws.cell(column=col+3, row=row,   value=dauer_2) #lat
                ws.cell(column=col+4, row=row,   value=distanz_3) #lat
                ws.cell(column=col+5, row=row,   value=dauer_3) #lat
                
        row=row+1


wb.save(hauptdatei)
print("Fertig. Datei gespeichert.")











